# course marks
# kieran hulsman

# openpyxl import
from openpyxl import load_workbook

def col_sum (ws, col):

    added = 0
    for row in range(3, 17): # range of worksheet
        cell = ws['{}{}'.format(col, row)].value # cell value

        try:
            added = added + cell
        except TypeError: # prevents operating with empty cells
            pass
    
    return added # returns col sum

# sub avg calculator
def sub_avg (ws, received_col, total_col, weight):

    # received
    received = col_sum(ws, received_col)

    # total
    total = col_sum(ws, total_col)

    return (received / total) * weight # returns weighted sub avg

    # standard weighted avg
def standard_avg (course_code, ku, ti, c, a): # pptx sheets are named after course code

    ws = course_marks[course_code]

    # cols: A, B
    ku = sub_avg(ws, 'A', 'B', ku)

    # cols: C, D
    ti = sub_avg(ws, 'C', 'D', ti)

    # cols: E, F
    c = sub_avg(ws, 'E', 'F', c)

    # cols: G, H
    a = sub_avg(ws, 'G', 'H', a)

    return (ku + ti + c + a) # returns avg

def display_avg(name, avg):
    print('{}: {} %'.format(name, round(avg, 2))) # displays avg to 2 decimal places
    # don't need to return anything because it's just being printed to console

# determines grade
grade = int(input('\ngrade: '))

if grade == 11:
   
    # excel file activation
    path = '/Users/kieranhulsman/Documents/School/Mark Worksheets/Grade 11.xlsx'
    course_marks = load_workbook(path)


    # ICS3U
    ws = course_marks['ICS3U']

    ics3u = (ws['A2'].value) * 100 # computer science mark
    # openpyxl converts % format into decimal
    # *100 converts it back


    # SCH3U
    sch3u = standard_avg('SCH3U', ku=30, ti=30, c=20, a=20)

    # SPH3U
    sph3u = standard_avg('SPH3U', ku=30, ti=30, c=20, a=20) # currently N/A (dropping the mark appeal, happy with final override)

    # mccoubrey's weighting ovveride (from markbook updates)
    ws = course_marks['SPH3U']

    # will assign var the latested override value
    for i in range(2, 17):
        try:
            sph3u_override = (ws['n{}'.format(i)].value * 100)
        except TypeError:
            pass

    # ENG3U

    # weight factor marking scheme
    received_weight = 0 # counter var
    total_weight = 0 # counter var
    for i in range(3, 17):

        ws = course_marks['ENG3U'] # worksheet
        
        def value(col): # returns value of col at row i
            return ws['{}{}'.format(col, i)].value

        # individual categories for each assignment
        def assignment_cat (r_col, t_col, cat_weight):
            r_value = value(r_col) # received
            t_value = value(t_col) # total

            return ((r_value / t_value) * cat_weight) / 100

        # calculating weighted category marks for each assignment
        try:
            ku_mark = assignment_cat('A', 'B', 25) # ku
            ti_mark = assignment_cat('C', 'D', 25) # ti
            c_mark = assignment_cat('E', 'F', 25) # c
            a_mark = assignment_cat('G', 'H', 25) # a

            weight_factor = value('I') # col where the assignment weight factors are stored
            total_weight += weight_factor

            percent_received = (ku_mark + ti_mark + c_mark + a_mark) * weight_factor # percent received from each assignment
            received_weight += percent_received # course avg
        
        except TypeError: # if cell is empty
            pass

    eng3u = (received_weight / total_weight) * 100

    # courses in progress
    #print('\n---CURRENT---') # currently N/A (no current courses)

    # finished courses
    print('\n---FINISHED---')
    display_avg('ICS3U', ics3u) # computer science
    display_avg('SCH3U', sch3u) # chem
    display_avg('SPH3U', sph3u_override) # physics (with override)
    display_avg('ENG3U', eng3u) # english

    mcr3u = 96 # placeholder: goal mark, used in projected avg calcs
    '''
    display_avg('MCR3U', mcr3u) # functions
    '''

    # gr11 avg
    grade11_avg = (round(ics3u) + round(sch3u) + round(sph3u_override) + round(eng3u) + round(mcr3u) * 2) / 6

    print('\n---GRADE 11---') # gr11 header
    display_avg('Grade 11: ', grade11_avg) # grade 11 avg

    # win/loss calcs

    # goals
    ics3u_goal = 99
    sch3u_goal = 96
    sph3u_goal = 97
    eng3u_goal = 92
    mcr3u_goal = 96

    # individual course w/l
    ics3u_wl = round(ics3u) - ics3u_goal
    sch3u_wl = round(sch3u) - sch3u_goal
    sph3u_wl = round(sph3u_override) - sph3u_goal
    eng3u_wl = round(eng3u) - eng3u_goal
    mcr3u_wl = round(mcr3u) - mcr3u_goal

    gr11_wl = ics3u_wl + sch3u_wl + sph3u_wl + eng3u_wl # + (mcr3u_wl * 2)

    print('W/L: {}\n'.format(gr11_wl)) # displays win/loss

elif grade == 12:
    
    # excel file activation
    path = '/Users/kieranhulsman/Documents/School/Mark Worksheets/Grade 12.xlsx'
    course_marks = load_workbook(path)

    # ENG4U (english)
    eng4u = 98 # placeholder

    # ICS4U (computer science)
    ics4u = 99 # placeholder

    # MCV4U (calc)
    mcv4u = 96 # placeholder

    # MHF4U (functions)
    mhf4u = 96 # placeholder

    # SPH4U (physics)
    sph4u = 96 # placeholder

    # SCH4U (chem)
    sch4u = 97 # placeholder

    # courses in progress
    print('\n---CURRENT COURSES---') # current courses header
    display_avg('ENG4U', eng4u)
    display_avg('ICS4U', ics4u)

    # finished courses
    #print('\n---FINISHED COURSES---') # finished courses header

    # for when 2021-2022 school year starts
    '''
    display_avg('MCV4U', mcv4u)
    display_avg('MHF4U', mhf4u)
    display_avg('SPH4U', sph4u)
    display_avg('SCH4U', sch4u)
    '''

    # grade 12 avg
    print('\n---GRADE 12---') # gr12 header

    gr12_avg = (round(eng4u) + round(ics4u) + round(mcv4u) + round(mhf4u) + round(sph4u) + round(sch4u)) / 6
    display_avg('Top 6: ', gr12_avg)

    # win/loss calcs

    # goal avgs
    eng4u_goal = 98
    ics4u_goal = 99
    mcv4u_goal = 96
    mhf4u_goal = 96
    sph4u_goal = 96
    sch4u_goal = 97

    # individual course w/l
    eng4u_wl = round(eng4u) - eng4u_goal
    ics4u_wl = round(ics4u) - ics4u_goal
    mcv4u_wl = round(mcv4u) - mcv4u_goal
    mhf4u_wl = round(mhf4u) - mhf4u_goal
    sph4u_wl = round(sph4u) - sph4u_goal
    sch4u_wl = round(sch4u) - sch4u_goal

    # for when grade 12 courses start
    '''
    gr12_wl = eng4u_wl + ics4u_wl + mcv_wl + mph_wl + sph4u_wl + sch4u_wl
    print('W/L: {}\n'.format(gr12_wl))
    '''

else:
    print('input error: enter 11 or 12')