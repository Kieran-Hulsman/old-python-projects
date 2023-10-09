# kieran hulsman
# period table tool

# openpyxl library
from openpyxl import load_workbook
fn = 'ChemicalProperties.xlsx'
wb = load_workbook(filename= fn)
ws = wb['Sheet1']

# list of properties user can search for
properties_list = ['atomicnumber','group', 'period', 'atomicmass', 'electronegativity', 'ioniccharge', 'latinprefix', 'metallic', 'stateatsatp']

# list of colomns that correspond with searchable properties
property_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

# determines column (which property the user's searching for)
def col(query_value):

    # matches propety names to column letters
    property_name = {
        properties_list[0]: property_columns[0], # atomic number
        properties_list[1]: property_columns[1], # group
        properties_list[2]: property_columns[2], # period
        properties_list[3]: property_columns[3], # atomic mass
        properties_list[4]: property_columns[4], # electronegativity
        properties_list[5]: property_columns[5], # ionic charge
        properties_list[6]: property_columns[6], # latin prefix
        properties_list[7]: property_columns[7], # metallic
        properties_list[8]: property_columns[8]  # state at satp
    }

    return property_name.get(query_value) # returns colomn letter


# determines row
def row(query_value):

    # defines search range
    if len(query_value) > 2: # name
        search_range = ws['A2:A119']
    else: # symbol
        search_range = ws['B2:B119']

    # searches for element
    for row in search_range:
        for cell in row:
            if query_value == cell.value.lower():
                return cell.row # returns row

while(True): # repeats until user enters exit command

    # gets query
    query = input('enter query: ').lower().split('.')

    # checks if user requested to exit
    if len(query) == 1 and query[0] == 'exit': # user entered exit command
        break # ensures program doesn't output 'invalid query'

    try: # ValueError

        # gets element from user's query
        element = query[0]

        # displayes element name and symbol
        element_row = row(element)

        name = ws['A{}'.format(element_row)].value
        symbol = ws['B{}'.format(element_row)].value

        print('\n{} ({})'.format(name, symbol))

        if len(query) == 2 and query[1] == 'all': # user requested all properties
            
            # displays all properties
            for i in range(len(properties_list)):
                property_col = property_columns[i]
                property_row = row(element) 

                property_value = ws['{}{}'.format(property_col, property_row)].value
                print('\n\t{}: {}'.format(properties_list[i], property_value))

        else: # user requested a limited number of properties
            
            # displays elements searched property(ies)
            for i in range(1,len(query)):

                # determines property value
                property_col = col(query[i])
                property_row = row(element)

                property_value = ws['{}{}'.format(property_col, property_row)].value 
                print('\n\t{}: {}'.format(query[i], property_value))

    except ValueError: # user entered an invalid query
        print('invalid query')

    print('\n--------------------------------\n') # creates gap

wb.save(filename= fn)